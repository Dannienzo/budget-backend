from builtins import Exception, enumerate, range, str
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from django.template.loader import render_to_string
from rest_framework.permissions import IsAuthenticated
from rest_framework import status, viewsets, generics
from django.http import HttpResponse
from django.db import transaction as db_transaction
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth.models import User
from django.db import models
from django.db.models import Sum, Case, When, F, Q, Value, DecimalField, Count
from django.template.loader import render_to_string
from datetime import datetime, timedelta
import csv
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from io import StringIO
from decimal import Decimal
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models.functions import TruncMonth
from django.db import transaction as db_transaction
from django_filters.rest_framework import DjangoFilterBackend
from .serializers import (
    CategorySerializer,
    ExpenseSerializer,
    IncomeSerializer,
    BudgetSerializer,
    TransactionSerializer,
    CustomTokenObtainPairSerializer,
    UserProfileSerializer,
    AnalyticsSerializer,
    PasswordChangeSerializer,
)
from .models import Category, Expense, Income, Budget, Transaction, UserProfile


class DashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        total_income = (
            Transaction.objects.filter(user=user, type="income").aggregate(
                total=Sum("amount")
            )["total"]
            or 0
        )
        total_expense = (
            Transaction.objects.filter(user=user, type="expense").aggregate(
                total=Sum("amount")
            )["total"]
            or 0
        )

        try:
            total_income = float(total_income)
        except Exception:
            total_income = 0.0

        try:
            total_expense = float(total_expense)
        except Exception:
            total_expense = 0.0

        balance = total_income - total_expense
        latest_transactions = Transaction.objects.filter(user=user).order_by("-date")[
            :5
        ]

        data = {
            "username": user.username,
            "total_income": total_income,
            "total_expense": total_expense,
            "balance": balance,
            "recent_transactions": TransactionSerializer(
                latest_transactions, many=True
            ).data,
        }
        return Response(data)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user).order_by("-date")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class IncomeViewSet(viewsets.ModelViewSet):
    serializer_class = IncomeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user).order_by()

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class BudgetViewSet(viewsets.ModelViewSet):
    queryset = Budget.objects.all()
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Budget.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"])
    def summary(self, request):
        budgets = self.get_queryset()

        total_budget = budgets.aggregate(total=Sum("amount"))["total"] or 0
        total_spent = budgets.aggregate(total=Sum("spent"))["total"] or 0

        on_track = budgets.filter(spent__lte=models.F("amount")).count()
        over_budget = budgets.filter(spent__gt=models.F("amount")).count()
        count = budgets.count()

        return Response(
            {
                "total_budget": float(total_budget),
                "total_spent": float(total_spent),
                "total_remaining": float(total_budget - total_spent),
                "categories_count": count,
                "categories_on_track": on_track,
                "categories_over_budget": over_budget,
                "budget_adherence": (
                    (on_track / budgets.count() * 100) if budgets.count() > 0 else 0
                ),
            }
        )


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["date", "category", "amount"]

    def get_queryset(self):
        return Transaction.objects.filter(user=self.request.user).order_by("-date")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"])
    def recent(self, request):
        qs = self.get_queryset()[:5]
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"])
    def summary(self, request):
        try:
            qs = self.get_queryset()
            totals = qs.values("type").annotate(total=Sum("amount"))

            data = {"income": 0.0, "expense": 0.0}
            for t in totals:
                typ = t.get("type")
                total = t.get("total") or 0
                data[typ] = float(total)

            data["balance"] = data.get("income", 0.0) - data.get("expense", 0.0)
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


# class SpendingByCategoryView(APIView):
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         user = request.user
#         qs =(
#             Transaction.objects
#             .filter(user=user, type='expense', category__isnull=False)
#             .values('category__id', 'category__name')
#             .annotate(total=Sum('amount'))
#             .order_by('-total')
#         )

#         data = [
#             {
#                 "category_id": item['category__id'],
#                 "category": item['category__name'],
#                 "total": float(item['total']) if item['total'] is not None else 0.0
#             }
#             for item in qs
#         ]
#         return Response({"results": data})


# class IncomeVsExpenseByMonthView(APIView):
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         user = request.user

#         qs = (
#             Transaction.objects
#             .filter(user=user)
#             .annotate(month=TruncMonth('date'))
#             .values('month')
#             .annotate(
#                 income=Sum(Case(When(type='income', then=F('amount')), default=Value(0), output_field=DecimalField())),
#                 expense=Sum(Case(When(type='expense', then=F('amount')), default=Value(0), output_field=DecimalField()))
#             )
#             .order_by('month')
#         )

#         data = [
#             {
#                 "month": item['month'].strftime('%Y-%m') if item['month'] else None,
#                 "income": float(item['income'] or 0.0),
#                 "expense": float(item['expense'] or 0.0),
#                 "net": float((item['income'] or 0) - (item['expense'] or 0))
#             }
#             for item in qs
#         ]
#         return Response({"results": data})


# class MonthlySummaryView(APIView):
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         user = request.user
#         months = int(request.query_params.get('months', 12))

#         qs = (
#             Transaction.objects
#             .filter(user=user)
#             .annotate(month=TruncMonth('date'))
#             .values('month')
#             .annotate(
#                 income=Sum(Case(When(type='income', then=F('amount')), default=Value(0), output_field=DecimalField())),
#                 expense=Sum(Case(When(type='expense', then=F('amount')), default=Value(0), output_field=DecimalField()))
#             )
#             .order_by('month')
#         )

#         items = list(qs)
#         if months and len(items) > months:
#             items = items[-months:]

#         labels = []
#         income_series = []
#         expense_series = []
#         net_series = []

#         for it in items:
#             label = it['month'].strftime('%Y-%m') if it['month'] else None
#             inc = float(it['income'] or 0.0)
#             exp = float(it['expense'] or 0.0)
#             labels.append(label)
#             income_series.append(inc)
#             expense_series.append(exp)
#             net_series.append(inc - exp)

#         return Response({
#             "labels": labels,
#             "income": income_series,
#             "expense": expense_series,
#             "net": net_series
#         })


class AnalyticsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        user = request.user
        transactions = Transaction.objects.filter(user=user)

        total_income = transactions.filter(type="income").aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0")
        total_expenses = transactions.filter(type="expense").aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0")

        total_income_f = float(total_income)
        total_expenses_f = float(total_expenses)

        balance = total_income - total_expenses
        balance_f = float(balance)

        spending_rate = (
            (float(total_expenses) / float(total_income) * 100)
            if total_income != 0
            else 0.0
        )

        budgets = Budget.objects.filter(user=user)
        total_budget = budgets.aggregate(total=Sum("amount"))["total"] or Decimal("0")
        total_spent = budgets.aggregate(total=Sum("spent"))["total"] or Decimal("0")

        on_track = budgets.filter(spent__lte=models.F("amount")).count()
        budget_adherence = (
            (on_track / budgets.count() * 100) if budgets.count() > 0 else 0
        )

        data = {
            "total_income": total_income_f,
            "total_expenses": total_expenses_f,
            "balance": balance_f,
            "spending_rate": spending_rate,
            "budget_adherence": budget_adherence,
            "avg_monthly_expenses": float(total_expenses),
            "total_budget": float(total_budget),
            "total_spent": float(total_spent),
            "categories_on_track": on_track,
            "categories_over_budget": budgets.count() - on_track,
        }

        serializer = AnalyticsSerializer(data)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def monthly_trends(self, request):

        from .models import Transaction

        six_months_ago = datetime.now() - timedelta(days=180)
        transactions = Transaction.objects.filter(
            user=request.user, date__gte=six_months_ago
        )

        monthly_data = []
        for i in range(6):
            month_start = (datetime.now() - timedelta(days=30 * i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(
                days=1
            )

            month_transactions = transactions.filter(
                date__gte=month_start, date__lte=month_end
            )

            income = (
                month_transactions.filter(type="income").aggregate(total=Sum("amount"))[
                    "total"
                ]
                or 0
            )
            expenses = (
                month_transactions.filter(type="expense").aggregate(
                    total=Sum("amount")
                )["total"]
                or 0
            )

            monthly_data.append(
                {
                    "month": month_start.strftime("%b"),
                    "income": float(income),
                    "expenses": float(expenses),
                    "savings": float(income - expenses),
                }
            )
        return Response(monthly_data[::-1])


class UserProfileViewSet(viewsets.ModelViewSet):
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return UserProfile.objects.filter(user=self.request.user)

    def get_object(self):
        profile, created = UserProfile.objects.get_or_create(user=self.request.user)
        return profile

    def list(self, request):
        profile = self.get_object()
        serializer = self.get_serializer(profile)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        profile = self.get_object()
        serializer = self.get_serializer(profile, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        profile = self.get_object()
        serializer = self.get_serializer(profile, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=["get", "patch", "put"])
    def me(self, request):
        profile = self.get_object()

        if request.method == "GET":
            serializer = self.get_serializer(profile)
            return Response(serializer.data)

        elif request.method in ["PATCH", "PUT"]:
            serializer = self.get_serializer(profile, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="change-password")
    def change_password(self, request):
        """POST /api/profile/change-password/"""
        serializer = PasswordChangeSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {"message": "Password changed successfully"}, status=status.HTTP_200_OK
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_transactions_csv(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    transaction_type = request.GET.get("type")

    transactions = Transaction.objects.filter(user=request.user)

    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    if transaction_type and transaction_type != "all":
        transactions = transactions.filter(type=transaction_type)

    transactions = transactions.order_by("-date")

    response = HttpResponse(content_type="text/csv")
    filename = f'transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    writer.writerow(["Date", "Type", "Category", "Amount (₦)", "Description"])

    for txn in transactions:
        writer.writerow(
            [
                txn.date.strftime("%Y-%m-%d"),
                txn.type.capitalize(),
                txn.category.name if txn.category else "Uncategorized",
                f"{txn.amount:.2f}",
                txn.description or "",
            ]
        )

    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_transactions_excel(request):

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    transaction_type = request.GET.get("type")

    transactions = Transaction.objects.filter(user=request.user)

    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    if transaction_type and transaction_type != "all":
        transactions = transactions.filter(type=transaction_type)

    transactions = transactions.order_by("-date")

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill(
        start_color="1E40AF", end_color="1E40AF", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)

    headers = ["Date", "Type", "Category", "Amount (₦)", "Description"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, txn in enumerate(transactions, start=2):
        ws.cell(row=row_num, column=1, value=txn.date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=2, value=txn.type.capitalize())
        ws.cell(
            row=row_num,
            column=3,
            value=txn.category.name if txn.category else "Uncategorized",
        )

        amount_cell = ws.cell(row=row_num, column=4, value=float(txn.amount))
        amount_cell.number_format = "#,##0.00"

        if txn.type == "income":
            amount_cell.font = Font(color="16A34A")  # Green
        else:
            amount_cell.font = Font(color="DC2626")  # Red

        ws.cell(row=row_num, column=5, value=txn.description or "")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40

    ws_summary = wb.create_sheet("Summary")

    total_income = (
        transactions.filter(type="income").aggregate(Sum("amount"))["amount__sum"] or 0
    )
    total_expense = (
        transactions.filter(type="expense").aggregate(Sum("amount"))["amount__sum"] or 0
    )
    net_savings = total_income - total_expense

    ws_summary["A1"] = "Financial Summary"
    ws_summary["A1"].font = Font(bold=True, size=16, color="1E40AF")
    ws_summary.merge_cells("A1:B1")

    summary_data = [
        ("Total Income:", float(total_income)),
        ("Total Expenses:", float(total_expense)),
        ("Net Savings:", float(net_savings)),
    ]

    row = 3
    for label, value in summary_data:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        amount_cell = ws_summary.cell(row=row, column=2, value=value)
        amount_cell.number_format = "₦#,##0.00"

        # Color coding
        if "Income" in label:
            amount_cell.font = Font(color="16A34A", bold=True)
        elif "Expenses" in label:
            amount_cell.font = Font(color="DC2626", bold=True)
        elif "Savings" in label:
            color = "16A34A" if value >= 0 else "DC2626"
            amount_cell.font = Font(color=color, bold=True)

        row += 1

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20

    ws_categories = wb.create_sheet("By Category")

    # Header
    ws_categories["A1"] = "Expense by Category"
    ws_categories["A1"].font = Font(bold=True, size=16, color="1E40AF")
    ws_categories.merge_cells("A1:C1")

    # Column headers
    headers_cat = ["Category", "Amount (₦)", "Percentage"]
    for col_num, header in enumerate(headers_cat, start=1):
        cell = ws_categories.cell(row=3, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    expense_by_category = (
        transactions.filter(type="expense")
        .values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    row = 4
    for item in expense_by_category:
        category_name = item["category__name"] or "Uncategorized"
        amount = float(item["total"])
        percentage = (amount / float(total_expense) * 100) if total_expense > 0 else 0

        ws_categories.cell(row=row, column=1, value=category_name)

        amount_cell = ws_categories.cell(row=row, column=2, value=amount)
        amount_cell.number_format = "₦#,##0.00"

        percent_cell = ws_categories.cell(row=row, column=3, value=percentage / 100)
        percent_cell.number_format = "0.0%"

        row += 1

    ws_categories.column_dimensions["A"].width = 25
    ws_categories.column_dimensions["B"].width = 20
    ws_categories.column_dimensions["C"].width = 15

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f'transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)

    return response




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_monthly_pdf_report(request):
    """Generate a professional monthly PDF report"""
    
    # Get month and year from query params
    month = int(request.GET.get('month', datetime.now().month))
    year = int(request.GET.get('year', datetime.now().year))
    
    # Validate month
    if not (1 <= month <= 12):
        return Response({'error': 'Invalid month'}, status=400)
    
    # Calculate date range for the month
    start_date = datetime(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = datetime(year, month, last_day, 23, 59, 59)
    
    # Get transactions for the month
    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('category').order_by('-date')
    
    # Calculate summary statistics
    total_income = transactions.filter(type='income').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    total_expense = transactions.filter(type='expense').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    net_savings = float(total_income) - float(total_expense)
    savings_rate = (net_savings / float(total_income) * 100) if total_income > 0 else 0
    
    # Get top expense categories (FIXED - removed icon reference)
    expense_by_category = transactions.filter(type='expense').values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')[:5]
    
    # Convert to list and add percentages with default icons
    expense_categories = []
    # Map category names to icons (you can customize these)
    category_icons = {
        'Food': '🍔',
        'Transport': '🚗',
        'Entertainment': '🎬',
        'Shopping': '🛍️',
        'Bills': '💡',
        'Health': '🏥',
        'Education': '📚',
        'Other': '💰'
    }
    
    for item in expense_by_category:
        amount = float(item['total'])
        percentage = (amount / float(total_expense) * 100) if total_expense > 0 else 0
        category_name = item['category__name'] or 'Uncategorized'
        
        # Get icon from mapping or use default
        icon = category_icons.get(category_name, '💰')
        
        expense_categories.append({
            'name': category_name,
            'icon': icon,
            'amount': amount,
            'percentage': percentage
        })
    
    # Get top income sources (FIXED - removed icon reference)
    income_by_category = transactions.filter(type='income').values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')[:5]
    
    income_sources = []
    # Income category icons
    income_icons = {
        'Salary': '💵',
        'Freelance': '💼',
        'Investment': '📈',
        'Business': '🏢',
        'Gift': '🎁',
        'Other': '💰'
    }
    
    for item in income_by_category:
        amount = float(item['total'])
        percentage = (amount / float(total_income) * 100) if total_income > 0 else 0
        category_name = item['category__name'] or 'Uncategorized'
        
        # Get icon from mapping or use default
        icon = income_icons.get(category_name, '💵')
        
        income_sources.append({
            'name': category_name,
            'icon': icon,
            'amount': amount,
            'percentage': percentage
        })
    
    # Get daily spending trend (group by day)
    daily_expenses = transactions.filter(type='expense').extra(
        select={'day': 'date(date)'}
    ).values('day').annotate(
        total=Sum('amount')
    ).order_by('day')
    
    daily_trend = [
        {
            'day': item['day'].strftime('%d'),
            'amount': float(item['total'])
        }
        for item in daily_expenses
    ]
    
    recent_transactions = list(transactions[:10].values(
        'date', 'type', 'category__name', 'amount', 'description'
    ))
    
    for txn in recent_transactions:
        txn['date'] = txn['date'].strftime('%Y-%m-%d')
        txn['amount'] = float(txn['amount'])
        category_name = txn.pop('category__name') or 'Uncategorized'
        txn['category_name'] = category_name
        
        if txn['type'] == 'income':
            txn['category_icon'] = income_icons.get(category_name, '💵')
        else:
            txn['category_icon'] = category_icons.get(category_name, '💰')
    
    context = {
        'month_name': calendar.month_name[month],
        'year': year,
        'user': request.user,
        'total_income': float(total_income),
        'total_expense': float(total_expense),
        'net_savings': net_savings,
        'savings_rate': savings_rate,
        'expense_categories': expense_categories,
        'income_sources': income_sources,
        'daily_trend': daily_trend,
        'recent_transactions': recent_transactions,
        'transaction_count': transactions.count(),
        'generated_date': datetime.now().strftime('%B %d, %Y'),
    }
    
    html_string = render_to_string('reports/monthly_report.html', context)

    try:
        from weasyprint import HTML
        from weasyprint.text.fonts import FontConfiguration

        font_config = FontConfiguration()
        html = HTML(string=html_string)

        response = HttpResponse(content_type='application/pdf')
        filename = f'monthly_report_{year}_{month:02d}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
        html.write_pdf(response, font_config=font_config)
    
        return response
    
    except ImportError as e:
        return Response({
        'error': 'PDF generation is currently unavailable.',
        'detail': 'WeasyPrint dependencies are not properly installed.',
        'message': str(e)
    }, status=500)

    except Exception as e:
        return Response({
            'error': 'Failed to generate PDF',
            'detail': str(e)
        }, status=500)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_bank_csv(request):

    if 'file' not in request.FILES:
        return Response({'error': 'No file uploaded'}, status=400)
    
    csv_file = request.FILES['file']
    bank_type = request.data.get('bank_type', 'generic')

    if not csv_file.name.endswith('.csv'):
        return Response({'error': 'File must be a CSV'}, status=400)
    
    try:
        # Read CSV file
        file_data = csv_file.read().decode('utf-8')
        df = pd.read_csv(StringIO(file_data))

        df.columns = df.columns.str.strip()

        column_mappings = {
            'gtbank': {
                'date': 'Trans Date',
                'description': 'Narration',
                'debit': 'Debit',
                'credit': 'Credit',
                'balance': 'Balance'
            },
            'zenith': {
                'date': 'Transaction Date',
                'description': 'Narration',
                'debit': 'Debit',
                'credit': 'Credit',
                'balance': 'Balance'
            },
            'kuda': {
                'date': 'Date',
                'description': 'Description',
                'debit': 'Debit',
                'credit': 'Credit',
                'balance': 'Balance'
            },
            'access': {
                'date': 'Date',
                'description': 'Narration',
                'debit': 'Withdrawal',
                'credit': 'Deposit',
                'balance': 'Balance'
            },
            'generic': {
                'date': 'Date',
                'description': 'Description',
                'debit': 'Debit',
                'credit': 'Credit',
                'balance': 'Balance'
            }
        }

        mapping = column_mappings.get(bank_type, column_mappings["generic"])

        required_cols = [mapping['date'], mapping['description']]
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            return Response({
                'error': f'Missing required columns: {", ".join(missing_cols)}',
                'available_columns': list(df.columns),
                'bank_type': bank_type
            }, status=400)
        
        transactions_to_create = []
        duplicate_count = 0
        error_count = 0
        
        for index, row in df.iterrows():
            try:
                date_str = str(row[mapping["date"]]).strip()
                try:
                    for date_format in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                        try:
                            transaction_date= datetime.strptime(date_str, date_format).date()
                            break
                        except ValueError:
                            continue
                    else:
                        error_count += 1
                        continue
                except Exception:
                    error_count += 1
                    continue
                
                description= str(row[mapping["description"]]).strip()

                debit = row.get(mapping.get('debit'), 0)
                credit = row.get(mapping.get('credit'), 0)

                try:
                    debit_amount = float(str(debit).replace(',', '')) if pd.notna(debit) and str(debit).strip() else 0
                except (ValueError, AttributeError):
                    debit_amount = 0

                try:
                    credit_amount = float(str(credit).replace(',', '')) if pd.notna(credit) and str(credit).strip() else 0
                except (ValueError, AttributeError):
                    credit_amount = 0

                if credit-amount > 0:
                    txn_type = 'income'
                    amount = credit_amount
                else:
                    txn_type = 'expense'
                    amount = debit_amount


                existing = Transaction.objects.filter(
                    user=request.user,
                    date=transaction_date,
                    amount=amount,
                    description__icontains=description[:50]  # Check first 50 chars
                ).exists()

                if existing:
                    duplicate_count += 1
                    continue

                category = None
                description_lower = description.lower()

                if any(word in description_lower for word in ['transfer', 'salary', 'payment received']):
                    category_name = 'Salary' if txn_type == 'income' else 'Other'
                elif any(word in description_lower for word in ['atm', 'withdrawal', 'pos']):
                    category_name = 'Shopping'
                elif any(word in description_lower for word in ['uber', 'bolt', 'transport', 'fuel']):
                    category_name = 'Transport'
                elif any(word in description_lower for word in ['restaurant', 'food', 'kfc', 'dominos']):
                    category_name = 'Food'
                elif any(word in description_lower for word in ['netflix', 'spotify', 'dstv', 'showmax']):
                    category_name = 'Entertainment'
                elif any(word in description_lower for word in ['electricity', 'water', 'nepa', 'phcn']):
                    category_name = 'Bills'
                else:
                    category_name = 'Other'


                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'description': f'Auto-created from CSV import'}
                )

                transactions_to_create.append(Transaction(
                    user=request.user,
                    type=txn_type,
                    category=category,
                    amount=amount,
                    description=description,
                    date=transaction_date
                ))

            except Exception as e:
                error_count += 1
                print(f"Error processing row {index}: {str(e)}")
                continue
        
        created_count = 0
        if transactions_to_create:
            with db_transaction.atomic():
                Transaction.objects.bulk_create(transactions_to_create)
                created_count = len(transactions_to_create)

        return Response({
            'success': True,
            'created': created_count,
            'duplicates': duplicate_count,
            'errors': error_count,
            'total_rows': len(df),
            'message': f'Successfully imported {created_count} transactions'
        }, status=201)
    
    except pd.errors.EmptyDataError:
        return  Response({'error': 'CSV file is empty'}, status=400)
    except Exception as e:
        return Response({
            'error': f'Error processing CSV: {str(e)}',
            'detail': str(e)
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def preview_bank_csv(request):
    """
    Preview CSV file before importing
    Returns first 10 rows and detected columns
    """
    
    if 'file' not in request.FILES:
        return Response({'error': 'No file uploaded'}, status=400)
    
    csv_file = request.FILES['file']
    
    if not csv_file.name.endswith('.csv'):
        return Response({'error': 'File must be a CSV'}, status=400)
    
    try:
        # Read CSV file
        file_data = csv_file.read().decode('utf-8')
        df = pd.read_csv(StringIO(file_data))
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Get first 10 rows
        preview_data = df.head(10).to_dict('records')
        
        # Convert any NaN to None for JSON serialization
        for row in preview_data:
            for key, value in row.items():
                if pd.isna(value):
                    row[key] = None
        
        return Response({
            'columns': list(df.columns),
            'row_count': len(df),
            'preview': preview_data,
            'detected_bank': detect_bank_type(df.columns)
        }, status=200)
        
    except Exception as e:
        return Response({
            'error': f'Error reading CSV: {str(e)}'
        }, status=500)


def detect_bank_type(columns):
    """
    Auto-detect bank type based on column names
    """
    columns_lower = [col.lower() for col in columns]
    
    if 'trans date' in columns_lower:
        return 'gtbank'
    elif 'transaction date' in columns_lower:
        return 'zenith'
    elif 'withdrawal' in columns_lower and 'deposit' in columns_lower:
        return 'access'
    elif any('kuda' in col for col in columns_lower):
        return 'kuda'
    else:
        return 'generic'